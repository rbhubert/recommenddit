import os.path

import openpyxl

from information_recovery.reddit_connection import collect_submissions

file_name = "reddit_info.xlsx"
exist_file = os.path.exists(file_name)

if not exist_file:
    workbook = openpyxl.Workbook(file_name)
    workbook.create_sheet("subreddits", 0)
    workbook.create_sheet("submissions", 1)
    workbook.create_sheet("comments", 2)
    workbook.create_sheet("crossposts", 3)
    workbook.save(file_name)
    workbook.close()


def add_subreddit(subreddit):
    workbook = openpyxl.load_workbook(file_name)
    subreddits_worksheet = workbook["subreddits"]
    entry = (subreddit.name, subreddit.description, subreddit.date_created, subreddit.nsfw, subreddit.subscribers)
    subreddits_worksheet.append(entry)

    workbook.save(file_name)
    workbook.close()


def add_submission(submission):
    workbook = openpyxl.load_workbook(file_name)
    submission_worksheet = workbook["submissions"]
    entry = (submission.id, submission.title, submission.author, submission.date_created, submission.nsfw,
             submission.post_type, submission.upvote_ratio, submission.total_awards, submission.num_crossposts,
             submission.text, submission.video_duration, submission.category, submission.subreddit)
    submission_worksheet.append(entry)

    workbook.save(file_name)
    workbook.close()


def add_comment(comment):
    workbook = openpyxl.load_workbook(file_name)
    comments_worksheet = workbook["comments"]
    entry = (comment.id, comment.text, comment.author, comment.date_created, comment.parent_id,
             comment.submission_id, comment.upvote_ratio, comment.pinned)
    comments_worksheet.append(entry)

    workbook.save(file_name)
    workbook.close()


def add_crosspost(crosspost):
    workbook = openpyxl.load_workbook(file_name)
    crossposts_worksheet = workbook["crossposts"]
    entry = (crosspost.crosspost_parent_id, crosspost.post_id)
    crossposts_worksheet.append(entry)

    workbook.save(file_name)
    workbook.close()


def collect_subreddits(subreddits: [str]):
    """
    Go through the list of subreddits collecting all the submissions, crossposts and comments, and them save them
    in an excel (in batch).
    :param subreddits: list of str representing subreddits. Can be null.
    """

    for subreddit in subreddits:
        print(f"Getting posts from subreddit: '{subreddit}'.")

        subreddit_info, submissions, crossposts = collect_submissions(subreddit)

        # Update Excel
        add_subreddit(subreddit_info)
        for subm in submissions:
            add_submission(subm)
            for comm in subm.comments:
                add_comment(comm)

        for crossp in crossposts:
            add_crosspost(crossp)

        print(f"\t Saving information of subreddit: '{subreddit}'.")
